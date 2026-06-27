"""Construção/edição determinística de ficheiros .xlsx via openpyxl (zero alucinação).

O LLM produz apenas a *spec* (JSON); aqui geramos o binário real — fórmulas
verdadeiras, cabeçalho a negrito, freeze panes, larguras e formatos numéricos.
"""

from __future__ import annotations

import base64
import io
import logging
from typing import Any

_logger = logging.getLogger(__name__)

_MAX_XLSX_BYTES = 8 * 1024 * 1024
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class XlsxBuildError(RuntimeError):
    """openpyxl ausente ou falha irrecuperável ao construir o workbook."""


def _require_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover - depende do ambiente
        raise XlsxBuildError("openpyxl não instalado — necessário para gerar .xlsx.") from exc
    return openpyxl


def _style_header(ws: Any, ncols: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    font = Font(bold=True, color="FFFFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _auto_width(
    ws: Any, columns: list[str], rows: list[list[Any]], overrides: dict[str, float]
) -> None:
    from openpyxl.utils import get_column_letter

    ncols = max(len(columns), max((len(r) for r in rows), default=0))
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        if letter in overrides:
            ws.column_dimensions[letter].width = overrides[letter]
            continue
        max_len = len(str(columns[c - 1])) if c - 1 < len(columns) else 8
        for r in rows:
            if c - 1 < len(r) and r[c - 1] is not None:
                max_len = max(max_len, len(str(r[c - 1])))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def _write_sheet(ws: Any, sheet: dict[str, Any]) -> None:
    columns = sheet.get("columns") or []
    rows = sheet.get("rows") or []
    start_row = 1
    if columns:
        for c, header in enumerate(columns, start=1):
            ws.cell(row=1, column=c, value=header)
        _style_header(ws, len(columns))
        start_row = 2
    for r_off, row in enumerate(rows):
        for c, value in enumerate(row, start=1):
            ws.cell(row=start_row + r_off, column=c, value=value)
    # Formatos numéricos por coluna (letra -> formato).
    for letter, fmt in (sheet.get("number_formats") or {}).items():
        col_idx = _col_letter_to_index(letter)
        if col_idx is None:
            continue
        for r in range(start_row, start_row + len(rows)):
            ws.cell(row=r, column=col_idx).number_format = fmt
    _auto_width(ws, columns, rows, sheet.get("column_widths") or {})
    if columns and sheet.get("freeze_header", True):
        ws.freeze_panes = "A2"


def _col_letter_to_index(letter: str) -> int | None:
    s = str(letter).strip().upper()
    if not s.isalpha():
        return None
    idx = 0
    for ch in s:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def build_xlsx_bytes(spec: dict[str, Any]) -> bytes:
    openpyxl = _require_openpyxl()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = spec.get("sheets") or []
    if not sheets:
        sheets = [{"name": "Folha1", "columns": [], "rows": []}]
    for sheet in sheets:
        ws = wb.create_sheet(title=(sheet.get("name") or "Folha")[:31])
        _write_sheet(ws, sheet)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def load_workbook_from_bytes(data: bytes) -> Any:
    openpyxl = _require_openpyxl()
    return openpyxl.load_workbook(io.BytesIO(data), data_only=False)


def apply_edit_ops(data: bytes, plan: dict[str, Any]) -> bytes:
    """Carrega a planilha e aplica operações mínimas, preservando o restante."""
    wb = load_workbook_from_bytes(data)
    default_sheet = plan.get("target_sheet") or wb.sheetnames[0]

    def _ws(name: str | None) -> Any:
        title = name or default_sheet
        if title in wb.sheetnames:
            return wb[title]
        return wb[wb.sheetnames[0]]

    for op in plan.get("ops") or []:
        kind = op.get("op")
        ws = _ws(op.get("sheet"))
        if kind == "set_cell":
            ws[str(op["cell"])] = op.get("value")
        elif kind == "set_formula":
            formula = str(op.get("formula") or "")
            ws[str(op["cell"])] = formula if formula.startswith("=") else f"={formula}"
        elif kind == "add_rows":
            for row in op.get("rows") or []:
                ws.append(row)
        elif kind == "add_column":
            col = (ws.max_column or 0) + 1
            ws.cell(row=1, column=col, value=op.get("header"))
            for i, val in enumerate(op.get("values") or [], start=2):
                ws.cell(row=i, column=col, value=val)
        elif kind == "delete_row":
            try:
                ws.delete_rows(int(op.get("row")))
            except (TypeError, ValueError):
                continue
        elif kind == "rename_sheet":
            ws.title = str(op.get("to") or ws.title)[:31]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_xlsx_metadata(
    xlsx_bytes: bytes,
    *,
    filename: str,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    meta: dict[str, Any] = {
        "document_kind": "xlsx_result",
        "document_format": _XLSX_MIME,
        "xlsx_export_suggested_filename": filename,
    }
    if extra:
        meta.update(extra)
    if len(xlsx_bytes) <= _MAX_XLSX_BYTES:
        meta["xlsx_document_base64"] = base64.b64encode(xlsx_bytes).decode("ascii")
        meta["xlsx_size_bytes"] = len(xlsx_bytes)
    else:
        meta["xlsx_too_large"] = True
        meta["xlsx_size_bytes"] = len(xlsx_bytes)
    return meta


def ensure_xlsx_filename(name: str, *, default: str = "planilha.xlsx") -> str:
    n = (name or "").strip() or default
    if not n.lower().endswith(".xlsx"):
        n = n.rsplit(".", 1)[0] if "." in n else n
        n = f"{n}.xlsx"
    return n
