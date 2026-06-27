"""Leitura determinística (zero-token) de planilhas anexadas (.xlsx/.csv).

Produz um *digest* compacto (folhas, cabeçalhos, dimensões, amostra de linhas e
fórmulas) que cabe no budget de contexto do LLM — nunca despeja o ficheiro inteiro.
Degrada graciosamente se ``openpyxl`` não estiver instalado.
"""

from __future__ import annotations

import base64
import binascii
import csv
import io
import logging
from typing import Any

_logger = logging.getLogger(__name__)

# Limites do digest para respeitar o budget de tokens.
_MAX_SAMPLE_ROWS = 30
_MAX_COLS = 40
_MAX_CELL_CHARS = 200


def decode_attachment(att: dict[str, Any]) -> bytes | None:
    raw = str(att.get("data_base64") or "").strip()
    if not raw:
        return None
    if raw.startswith("data:") and "," in raw:
        raw = raw.split(",", 1)[1]
    try:
        return base64.b64decode(raw, validate=False)
    except (binascii.Error, ValueError):
        return None


def _is_csv(name: str, mime: str) -> bool:
    return name.lower().endswith(".csv") or "csv" in mime.lower()


def _clip_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    s = str(value)
    return s if len(s) <= _MAX_CELL_CHARS else s[: _MAX_CELL_CHARS - 1] + "…"


def _read_csv_digest(data: bytes, filename: str) -> dict[str, Any]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    headers = [_clip_cell(c) for c in rows[0][:_MAX_COLS]] if rows else []
    sample = [[_clip_cell(c) for c in r[:_MAX_COLS]] for r in rows[1 : _MAX_SAMPLE_ROWS + 1]]
    return {
        "filename": filename,
        "kind": "csv",
        "sheets": [
            {
                "name": "CSV",
                "max_row": len(rows),
                "max_col": max((len(r) for r in rows), default=0),
                "headers": headers,
                "rows": sample,
                "formulas": [],
            }
        ],
        "error": "",
    }


def _read_xlsx_digest(data: bytes, filename: str) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        return {
            "filename": filename,
            "kind": "xlsx",
            "sheets": [],
            "error": "openpyxl não instalado — instale a dependência para ler .xlsx.",
        }
    sheets: list[dict[str, Any]] = []
    try:
        # data_only=False mantém fórmulas; lemos valores e fórmulas separadamente.
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False, read_only=True)
        wb_vals = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        for ws in wb.worksheets:
            ws_vals = wb_vals[ws.title]
            max_row = ws.max_row or 0
            max_col = min(ws.max_column or 0, _MAX_COLS)
            headers: list[Any] = []
            rows: list[list[Any]] = []
            formulas: list[dict[str, Any]] = []
            for r_idx, row in enumerate(
                ws.iter_rows(
                    min_row=1, max_row=min(max_row, _MAX_SAMPLE_ROWS + 1), max_col=max_col
                ),
                start=1,
            ):
                vals_row = []
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        formulas.append({"cell": cell.coordinate, "formula": val})
                        computed = ws_vals[cell.coordinate].value
                        vals_row.append(_clip_cell(computed if computed is not None else val))
                    else:
                        vals_row.append(_clip_cell(val))
                if r_idx == 1:
                    headers = vals_row
                else:
                    rows.append(vals_row)
            sheets.append(
                {
                    "name": ws.title,
                    "max_row": max_row,
                    "max_col": ws.max_column or 0,
                    "headers": headers,
                    "rows": rows,
                    "formulas": formulas[:40],
                }
            )
        wb.close()
        wb_vals.close()
    except Exception as exc:  # noqa: BLE001
        return {
            "filename": filename,
            "kind": "xlsx",
            "sheets": sheets,
            "error": f"Falha ao abrir a planilha: {str(exc)[:200]}",
        }
    return {"filename": filename, "kind": "xlsx", "sheets": sheets, "error": ""}


def read_workbook_digest(data: bytes, *, filename: str, mime_type: str = "") -> dict[str, Any]:
    if _is_csv(filename, mime_type):
        try:
            return _read_csv_digest(data, filename)
        except Exception as exc:  # noqa: BLE001
            return {"filename": filename, "kind": "csv", "sheets": [], "error": str(exc)[:200]}
    return _read_xlsx_digest(data, filename)


def digest_to_markdown(digest: dict[str, Any]) -> str:
    """Serializa o digest para um bloco compacto para o prompt do LLM."""
    parts: list[str] = [f"# Planilha: {digest.get('filename')} ({digest.get('kind')})"]
    if digest.get("error"):
        parts.append(f"> Aviso: {digest['error']}")
    for sheet in digest.get("sheets") or []:
        parts.append(
            f"## Folha: {sheet.get('name')} "
            f"({sheet.get('max_row', 0)} linhas x {sheet.get('max_col', 0)} colunas)"
        )
        headers = sheet.get("headers") or []
        if headers:
            parts.append("| " + " | ".join(str(h) for h in headers) + " |")
            parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for row in sheet.get("rows") or []:
            parts.append("| " + " | ".join(str(c) for c in row) + " |")
        formulas = sheet.get("formulas") or []
        if formulas:
            parts.append(
                "Fórmulas: " + "; ".join(f"{f['cell']}={f['formula']}" for f in formulas[:15])
            )
    return "\n".join(parts).strip()
