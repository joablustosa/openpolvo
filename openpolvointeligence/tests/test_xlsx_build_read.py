"""Build (openpyxl) e leitura/digest determinísticos de planilhas."""

from __future__ import annotations

import base64
import importlib.util

import pytest

from openpolvointeligence.graphs.xlsx_full.xlsx_build_logic import (
    apply_edit_ops,
    build_xlsx_bytes,
    build_xlsx_metadata,
    ensure_xlsx_filename,
)
from openpolvointeligence.graphs.xlsx_full.xlsx_read_logic import (
    decode_attachment,
    digest_to_markdown,
    read_workbook_digest,
)
from openpolvointeligence.graphs.xlsx_full.xlsx_spec_logic import normalize_edit_plan

_HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None

_SPEC = {
    "filename": "vendas.xlsx",
    "sheets": [
        {
            "name": "Vendas",
            "columns": ["Produto", "Qtd", "Preço", "Total"],
            "rows": [
                ["A", 2, 10.0, "=B2*C2"],
                ["B", 3, 5.0, "=B3*C3"],
            ],
            "number_formats": {"C": "#,##0.00", "D": "#,##0.00"},
            "freeze_header": True,
        }
    ],
}


def test_decode_attachment_roundtrip() -> None:
    data = base64.b64encode(b"col1,col2\n1,2\n").decode()
    assert decode_attachment({"data_base64": data}) == b"col1,col2\n1,2\n"


def test_read_csv_digest() -> None:
    csv_bytes = b"Nome,Idade\nAna,30\nJoao,25\n"
    digest = read_workbook_digest(csv_bytes, filename="pessoas.csv", mime_type="text/csv")
    assert digest["kind"] == "csv"
    sheet = digest["sheets"][0]
    assert sheet["headers"] == ["Nome", "Idade"]
    assert sheet["rows"][0] == ["Ana", "30"]


def test_digest_to_markdown_includes_header() -> None:
    csv_bytes = b"A,B\n1,2\n"
    md = digest_to_markdown(read_workbook_digest(csv_bytes, filename="t.csv", mime_type="text/csv"))
    assert "| A | B |" in md


def test_ensure_xlsx_filename() -> None:
    assert ensure_xlsx_filename("orc") == "orc.xlsx"
    assert ensure_xlsx_filename("orc.csv") == "orc.xlsx"
    assert ensure_xlsx_filename("") == "planilha.xlsx"


def test_build_xlsx_metadata_embeds_base64() -> None:
    meta = build_xlsx_metadata(b"binary-bytes", filename="x.xlsx")
    assert meta["document_kind"] == "xlsx_result"
    assert meta["xlsx_export_suggested_filename"] == "x.xlsx"
    assert base64.b64decode(meta["xlsx_document_base64"]) == b"binary-bytes"
    assert meta["xlsx_size_bytes"] == len(b"binary-bytes")


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="requer openpyxl instalado")
def test_build_xlsx_bytes_and_read_back() -> None:
    import io

    import openpyxl  # type: ignore[import-untyped]

    data = build_xlsx_bytes(_SPEC)
    assert data[:2] == b"PK"  # zip/xlsx magic

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Vendas"]
    assert ws["A1"].value == "Produto"
    assert ws["A2"].value == "A"
    assert ws["D2"].value == "=B2*C2"
    assert ws.freeze_panes == "A2"


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="requer openpyxl instalado")
def test_read_xlsx_digest_roundtrip() -> None:
    data = build_xlsx_bytes(_SPEC)
    digest = read_workbook_digest(data, filename="vendas.xlsx")
    assert digest["kind"] == "xlsx"
    sheet = digest["sheets"][0]
    assert sheet["name"] == "Vendas"
    assert sheet["headers"] == ["Produto", "Qtd", "Preço", "Total"]
    assert any(f["formula"] == "=B2*C2" for f in sheet["formulas"])


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="requer openpyxl instalado")
def test_apply_edit_ops_set_cell_and_formula() -> None:
    import io

    import openpyxl  # type: ignore[import-untyped]

    data = build_xlsx_bytes(_SPEC)
    plan = normalize_edit_plan(
        {
            "target_sheet": "Vendas",
            "ops": [
                {"op": "set_cell", "sheet": "Vendas", "cell": "B2", "value": 9},
                {"op": "set_formula", "sheet": "Vendas", "cell": "E2", "formula": "B2+C2"},
                {"op": "add_rows", "sheet": "Vendas", "rows": [["C", 1, 1.0, "=B4*C4"]]},
            ],
        }
    )
    edited = apply_edit_ops(data, plan)
    wb = openpyxl.load_workbook(io.BytesIO(edited))
    ws = wb["Vendas"]
    assert ws["B2"].value == 9
    assert ws["E2"].value == "=B2+C2"
    assert ws["A4"].value == "C"
